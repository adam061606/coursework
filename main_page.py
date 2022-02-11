import openpyxl
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import font, colorchooser, filedialog, messagebox
from tkinter.filedialog import askopenfilename, asksaveasfilename
from tkinter import ttk
from tkinter import*
import os

wb = load_workbook ('notepad database 2.xlsx')
ws = wb.active
print(ws['A5'].value)

row_count = ws.max_row
column_count = ws.max_column

print(row_count)
print(column_count)

root = tk.Tk()
root.title("My Notes")
root.geometry('1200x700')

text_edit = tk.Text(root)
text_edit.grid(row=0, column=1, sticky="nsew")

frame_button = tk.Frame(root, relief=tk.RAISED, bd=3)
frame_button.grid(row=0, column=0, sticky="ns")

#settings changing function, it will edit things on the database
def settings_change(): #done by kim
    title, subject, level, privacy, text, description = '', '', '', '', '', ''
    desclist, titlelist = [], []
    noteinfo = tk.Tk()
    noteinfo.wm_title("Note Information")
    noteinfo.geometry('700x300')

    title_lbl = ttk.Label(noteinfo, text= "Title")
    entry_title = ttk.Entry(noteinfo, width=30)
    title_lbl.grid(row=0, column=0)
    entry_title.grid(row=0, column=1)
    def title_done(): #this is to signify whether person has made a title, title cannot be empty
        title = entry_title.get()
        titlelist.clear()
        titlelist.append(title)
    title_button = ttk.Button(noteinfo, text="Done", command=title_done)
    title_button.grid(row=0, column=2)
    
    desc_lbl = ttk.Label(noteinfo,text ='New Description:')
    entry_desc = ttk.Entry(noteinfo,width=30)
    desc_lbl.grid(row=1, column=0)
    entry_desc.grid(row=1, column=1)
    
    def desc_done(): #this ensures that there is a description entered, no description will cause error (no error message yet)
        description = entry_desc.get()
        desclist.clear()
        desclist.append(description)
    desc_button = ttk.Button(noteinfo, text="Done", command=desc_done)
    desc_button.grid(row=1, column=2)

    level_options = ["p1","p2", "p3", "p4", "p5", "p6", "s1", "s2", "s3","s4"] #dropdown options
    clickedL = StringVar()
    clickedL.set( "s4" )
    dropL = ttk.OptionMenu( noteinfo , clickedL , *level_options )
    dropL.grid(row=2, column=1, padx=20)
            
    subj_options = ["geography", "math", "chinese", "english", "history", "physics", "computing"] #dropdown options
    clicked = StringVar()
    clicked.set( "english" )
    drop = ttk.OptionMenu( noteinfo , clicked , *subj_options )
    drop.grid(row=3, column=1, padx=20)

    def subj_done(): #kinda useless right now
        subject = clicked.get()
    def level_done(): #kinda useless right now
        level = clickedL.get()

    level_button = ttk.Button(noteinfo, text="Done", command=subj_done) #kinda useless right now
    level_button.grid(row=3, column=2)
            
    subj_button = ttk.Button(noteinfo, text="Done", command=level_done) #kinda useless right now
    subj_button.grid(row=2, column=2)

    privacy_opt = ['','private', 'public'] #another dropdown
    clickedP = StringVar()
    clickedP.set("")
    dropP = ttk.OptionMenu( noteinfo , clickedP , *privacy_opt )
    dropP.grid(row=4, column=1, padx=20)

    def privacy_done(): #kinda useless right now
        privacy = clickedP.get()
    privacy_button = ttk.Button(noteinfo, text="Done", command = privacy_done)
    privacy_button.grid(row=4, column=2)

    #labels for the note information page, so that people know what the corresponding dropdowns are for
    subj_text = ttk.Label(noteinfo, text="Subject:")
    level_text = ttk.Label(noteinfo, text="Level:")
    privacy_text = ttk.Label(noteinfo, text="Privacy:")
    subj_text.grid(row=3, column=0)
    level_text.grid(row=2, column=0)
    privacy_text.grid(row=4, column=0)

    filename, favourite = '', 'not liked' #defaults

    def done(): #ensures that all the info is saved to database
        noteinfo.destroy()
        wb = load_workbook ('notepad database 2.xlsx')
        ws = wb.active
        ws.append([titlelist[0], clicked.get(), clickedL.get(), clickedP.get(),favourite, text, filename, desclist[0]]) #adds note info to database
        wb.save('notepad database 2.xlsx')

    done_button = ttk.Button(noteinfo, text="Done", command=done)
    done_button.grid(row=5, column=2)
    success= "your file has been uploaded" #sucess message
    text_edit.insert(tk.END, success)

    noteinfo.mainloop()


########################################################

def notepad(): #done by kim
    #the notepad screen
    main_screen_window = tk.Tk()
    main_screen_window.geometry('1200x800')
    main_screen_window.title('Notepad')

    tool_bar = ttk.Label(main_screen_window)
    tool_bar.pack(side=tk.TOP,fill=tk.X)

    filename_list = []


    font_tuple = tk.font.families()
    font_family = tk.StringVar()
    font_group = ttk.Combobox(tool_bar, width=30, textvariable=font_family, state='readonly' )
    font_group['values'] = font_tuple
    font_group.current(font_tuple.index('Arial'))
    font_group.grid(row=0, column=0, padx=5)


    size_variable = tk.IntVar()
    font_size=ttk.Combobox(tool_bar, width=14, textvariable = size_variable,state='readonly')
    font_size['values']=tuple(range(8,80,2))
    font_size.current(4)
    font_size.grid(row=0,column=1,padx=5)

    #the tool bar, so there's all the font stuff
    bold_button = ttk.Button(tool_bar, text= "bold")
    bold_button.grid(row=0, column=2, padx=5)

    italized_button = ttk.Button(tool_bar, text="italic")
    italized_button.grid(row=0, column=3,padx=5)

    underline_button=ttk.Button(tool_bar,text="underline")
    underline_button.grid(row=0, column=4,padx=5)

    font_color_button = ttk.Button(tool_bar, text="fontcolor")
    font_color_button.grid(row=0, column=5, padx=5)

    left_align_button = ttk.Button(tool_bar, text='left align')
    left_align_button.grid(row=0, column=6, padx=5)

    center_align_button = ttk.Button(tool_bar, text='center align')
    center_align_button.grid(row=0, column=7, padx=5)

    right_align_button = ttk.Button(tool_bar, text='right align')
    right_align_button.grid(row=0, column=8, padx=5)

    #this is the part where you write notes
    text_editor = tk.Text(main_screen_window)
    text_editor.config(wrap='word', relief=tk.FLAT)


    scroll_bar = tk.Scrollbar(main_screen_window)
    text_editor.focus_set()
    scroll_bar.pack(side=tk.RIGHT, fill=tk.Y)
    text_editor.pack(fill=tk.BOTH, expand=True)
    scroll_bar.config(command=text_editor.yview)
    text_editor.config(yscrollcommand=scroll_bar.set)


    present_font_family = 'Arial'
    present_font_size = 12

    def palitan_font(event=None):
        global present_font_family
        present_font_family = font_family.get()
        text_editor.config(font=(present_font_family, present_font_size))


    def palitan_size(event=None):
        global present_font_size
        present_font_size = size_variable.get()
        text_editor.config(font=(present_font_family, present_font_size))

    font_group.bind("<<ComboboxSelected>>", palitan_font)
    font_size.bind("<<ComboboxSelected>>", palitan_size)


    def palitan_bold():
        text_property = tk.font.Font(font=text_editor['font'])

        if text_property.actual()['weight']=='normal':
            text_editor.configure(font=(present_font_family,present_font_size,'bold'))
        if text_property.actual()['weight']=='bold':
            text_editor.configure(font=(present_font_family,present_font_size,'normal'))

    bold_button.configure(command=palitan_bold)


    def palitan_italized():
        text_property=tk.font.Font(font=text_editor['font'])

        if text_property.actual()['slant']=='roman' :
            text_editor.configure(font=(present_font_family, present_font_size,'italic'))
        if text_property.actual()['slant']=='italic' :
            text_editor.configure(font=(present_font_family, present_font_size,'normal'))

    italized_button.configure(command=palitan_italized)


    def underline():
        text_property=tk.font.Font(font=text_editor['font'])
    # upper line gives a dictionary whose attributes we are changing
        if text_property.actual()['underline']==0:
            text_editor.configure(font=(present_font_family, present_font_size,'underline'))
        if text_property.actual()['underline']==1:
            text_editor.configure(font=(present_font_family, present_font_size,'normal'))

    underline_button.configure(command=underline)

    def palitan_font_color():
        color_var = tk.colorchooser.askcolor()
        text_editor.configure(fg=color_var[1])

    font_color_button.configure(command=palitan_font_color)

    def left_align():
        text_content = text_editor.get(1.0, 'end')
        text_editor.tag_config('left', justify=tk.LEFT)
        text_editor.delete(1.0, tk.END)
        text_editor.insert(tk.INSERT,text_content,'left')

    left_align_button.configure(command=left_align)


    def center_align():
        text_content = text_editor.get(1.0, 'end')
        text_editor.tag_config('center',justify=tk.CENTER)
        text_editor.delete(1.0,tk.END)
        text_editor.insert(tk.INSERT,text_content,'center')

    center_align_button.configure(command=center_align)


    def right_align():
        text_content = text_editor.get(1.0, 'end')
        text_editor.tag_config('right',justify=tk.RIGHT)
        text_editor.delete(1.0,tk.END)
        text_editor.insert(tk.INSERT,text_content,'right')

    right_align_button.configure(command=right_align)



    text_editor.configure(font=('Arial',12))

    status_bar = ttk.Label(main_screen_window, text ='Status Bar')
    status_bar.pack(side=tk.BOTTOM)

    text_changed = False

    def changed(event=None):
        global text_changed
        if text_editor.edit_modified():###checks if any character is added or not
            text_changed= True
            words = len(text_editor.get(1.0, 'end-1c').split()) ##it even counts new line character so end-1c subtracts one char
            characters = len(text_editor.get(1.0,'end-1c'))
            status_bar.config(text=f' Words: {words} Characters : {characters}')
        text_editor.edit_modified(False)
    text_editor.bind('<<Modified>>',changed)


    url = ''

    def new_file_logo(event=None):
        global url
        settings_change()
        url = ''
        text_editor.delete(1.0,tk.END)
    #file.add_command(label='New' ,compound=tk.LEFT, accelerator ='Ctrl+N',command=new_file_logo)
    nfile = ttk.Button(tool_bar, text='New File', command=new_file_logo)
    nfile.grid(row=2, column=0, padx=5, pady=5)


    def open_file_logo(event=None):
        global url
        url = filedialog.askopenfilename(initialdir=os.getcwd(), title='Select File', filetypes=(('Text File', '*.txt'), ('All files', '*.*')))
        filename_list.clear()
        filename_list.append(url)
        try:
            with open(url, 'r') as fr:
                text_editor.delete(1.0, tk.END)
                text_editor.insert(1.0, fr.read())
        except FileNotFoundError:
            return
        except :
            return
        main_screen_window.title(os.path.basename(url))
    ofile = ttk.Button(tool_bar, text='Open File', command=open_file_logo)
    ofile.grid(row=2, column=1, padx=5, pady=5)

    def save_file_logo(event=None):
        global url
        try:
            if url:
                content = str(text_editor.get(1.0,tk.END))
                with open(url, 'w', encoding='utf-8') as fw:
                    fw.write(content)
            else :
                url = filedialog.asksaveasfile(mode='w', defaultextension='.txt', filetypes=(('Text File', '*.txt'),('All files','*.*')))
                content = text_editor.get(1.0, tk.END)
                url.write(content)
                url.close()
        except :
            return
        settings_change()
    sfile = ttk.Button(tool_bar, text='Save', command=save_file_logo)
    sfile.grid(row=2, column=3, padx=5, pady=5)

    def save_as_logo_file(event=None):
        global url
        try:
            content=text_editor.get(1.0,tk.END)
            url = filedialog.asksaveasfile(mode='w', defaultextension='.txt', filetypes=(('Text File', '*.txt'), ('All files','*.*')))
            url.write(content)
            url.close()
        except :
            return
        settings_change()
    safile = ttk.Button(tool_bar, text='Save As', command=save_as_logo_file)
    safile.grid(row=2, column=4, padx=5, pady=5)


    def exit_module(event=None):
        global url, text_changed
        try:
            if text_changed:
                boxmessage = messagebox.askyesnocancel('Warning','Do you want to save the file')
                if boxmessage is True :

                    if url:
                        content = text_editor.get(1.0,tk.END)
                        with open(url,'w',encoding='utf-8') as fw:
                            fw.write(content)
                            main_screen_window.destroy()
                    else:
                        content2 = str(text_editor.get(1.0,tk.END))
                        url = filedialog.asksaveasfile(mode='w',defaultextension='.txt',filetypes=(('Text File','*.txt'),('All files','*.*')))
                        url.write(content2)
                        url.close()
                        main_screen_window.destroy()
                elif boxmessage is False:
                    main_screen_window.destroy()
            else:
                main_screen_window.destroy()
        except:
            return


    efile = ttk.Button(tool_bar, text='Exit', command=exit_module)
    efile.grid(row=2, column=5, padx=5, pady=5)

    #sample data
    description = "the description for this note has not been updaed"
    subject, level = "subject not set", "level not set"

    def settings():
        settings_popup = tk.Tk()
        settings_popup.geometry('600x600')
        settings_popup.wm_title('Settings')

        settings_frame = tk.Frame(settings_popup, relief=tk.RAISED, bd=3)
        settings_frame.grid(row=0, column=0, sticky="ns")
        title, desc, subj, level = '', '', '', ''
        for i in range(row_count):
            H_cell = "H" + str(i)

            if filename_list[0] == H_cell:
                title, desc, subj, level = ("A" + str(i)), ("G" + str(i)), ("B"+str(i)), ("C"+str(i))
                pass
                
            title_txt =  ttk.Label(settings_popup, text="Title: {}".format(title))
            desc_txt =  ttk.Label(settings_popup, text="Description: {}".format(desc))
            subj_txt = ttk.Label(settings_popup, text="Subject: {}".format(subj))
            level_txt = ttk.Label(settings_popup, text="Level: {}".format(level))

            title_txt.grid(row=0, column=1)
            desc_txt.grid(row=1, column=1)
            subj_txt.grid(row=2, column=1)
            level_txt.grid(row=3, column=1)

        change_button = ttk.Button(settings_frame, text='modify settings', command=settings_change)
        change_button.grid(row=0, column=0)
    settings_button = ttk.Button(tool_bar, text="Settings", command=settings)
    settings_button.grid(row=2, column=6, padx=5)


    def search_module(event=None):

        def find():
            word = entry_find.get()
            text_editor.tag_remove('match','1.0',tk.END)
            matches = 0
            if word:
                start_pos = '1.0'
                while True:
                    start_pos = text_editor.search(word, start_pos, stopindex=tk.END)
                    if(not start_pos):
                        break
                    end_pos = f'{start_pos}+{len(word)}c'
                    text_editor.tag_add('match',start_pos,end_pos)
                    matches +=1
                    start_pos=end_pos
                    text_editor.tag_config('match',foreground='red',background='')


        def replace_module():
            word = entry_find.get()
            replace_module_text = entry_replace.get()
            content = text_editor.get(1.0,tk.END)
            new_content = content.replace(word,replace_module_text)
            text_editor.delete(1.0,tk.END)
            text_editor.insert(1.0,new_content)


        find_dlg = tk.Toplevel()
        find_dlg.geometry('450x250+500+200')
        find_dlg.resizable(0,0)

        #  frame
        find_frm = ttk.LabelFrame(find_dlg, text ='Find/Replace')
        find_frm.pack(pady=20)

        # labels
        find_txt_lbl = ttk.Label(find_frm,text ='Find :')
        replace_txt_lbl = ttk.Label(find_frm,text ='Replace')

        # entry boxes
        entry_find = ttk.Entry(find_frm,width=30)
        entry_replace = ttk.Entry(find_frm,width=30)


        # Button
        btn_find = ttk.Button(find_frm,text ='Find',command=find)
        btn_replace = ttk.Button(find_frm,text ='Replace',command=replace_module)

        # label grid
        find_txt_lbl.grid(row=0,column=0,padx=4,pady=4)
        replace_txt_lbl.grid(row=1,column=0,padx=4,pady=4)

        #entry grid
        entry_find.grid(row=0, column=1,padx=4,pady=4)
        entry_replace.grid(row=1, column=1,padx=4,pady=4)

        ##button grid
        btn_find.grid(row=2 ,column=0 ,padx=8,pady=4)
        btn_replace.grid(row=2 ,column=1 ,padx=8,pady=4)

        find_dlg.mainloop()

    findreplace = ttk.Button(tool_bar, text='Search/Replace', command=search_module)
    findreplace.grid(row=2, column=7, padx=5, pady=5)

    def favourite():
        like_msg = messagebox.askyesnocancel('Favourite', 'Do you want to add note to favourites?')
        for row in range(2,row_count):
            file_cell = "H" + str(row)
            if file_cell == filename: #find row of note
                cell = "E" +  str(row)
                if like_msg == True: ws[cell] = 'liked' #change the value in the favourite column on the database
                elif liked_msg == False: ws[cell] = 'not liked'
    
    liked_button = ttk.Button(tool_bar, text = 'favourite', command=favourite)
    liked_button.grid(row=2, column=8, padx=5, pady=5)
    
    main_screen_window.mainloop()

############################################################
button_upload = tk.Button(frame_button, text = "UPLOAD NOTES")
button_upload.grid(row=3, column=0, padx=5, pady=5)

def upload(): #done by kim
    title, subject, level, privacy, text, description = '', '', '', '', '', ''
    desclist, titlelist = [], []
    uploading_file = askopenfilename(filetypes=[("Text Files", "*.txt"), ["All files", "*.*"]]) #get location of file
    if not uploading_file:
        return
    text_edit.delete(1.0, tk.END)
    with open(uploading_file, "r") as file_input:
        text = file_input.read()

    filename = os.path.basename(uploading_file) #location of file will give the name of the file, this line is to remove the other stuff like location and users
    settings_change()
    success= "your file has been uploaded" #sucess message
    text_edit.insert(tk.END, success)

button_upload.config(command=upload)

def opening_file(): #done by kim
    file_location = askopenfilename(
        filetypes=[("Text Files", "*.txt"), ["All files", "*.*"]])
    if not file_location:
        return
    text_edit.delete(1.0, tk.END)
    with open(file_location, "r") as file_input:
        text = file_input.read()
        text_edit.insert(tk.END, text)
    root.title(f"my note - {file_location}")

def update_table(option_selected): #done by adam
    filtered_df = df[df.Team == option_selected]      
    return[filtered_df.to_dict('records')]

def pub1(): #redundant
    desc, subj, level, creator, created = pub1_info[0], pub1_info[1], pub1_info[2], pub1_info[3], pub1_info[4]
    note_info.config(text='Note Information:\n\nDescription:{}\nSubject & Level:{}, {}\nCreator:{}, {}'.format(desc, subj, level, creator, created))
    with open("sample_txt/biyuju.txt", "r") as file_input:
        text = file_input.read()
        text_edit.delete('1.0', END)
        text_edit.insert(tk.END, text)
        
for i in range (2,row_count+1): #done by adam
    a = "A" + str(i)
    d = "D" + str(i)
    print(a)
    if ws[d].value == "public":
        pub_1 = ttk.Button(root, text=ws[a].value, command=pub1)
        pub_1.grid(row=i+1, column=0)

#normal buttons
button_open = tk.Button(frame_button, text = "VIEW A NOTE", command=opening_file)
button_open.grid(row=0, column=0, padx=5, pady=5)

button_note = tk.Button(frame_button, text = "EDIT A NOTE", command=notepad)
button_note.grid(row=1, column=0, padx=5, pady=5)

button_liked = tk.Button(frame_button, text = "LIKED NOTES")
button_liked.grid(row=2, column=0, padx=5, pady=5)

button_upload = tk.Button(frame_button, text = "UPLOAD NOTES", command = upload)
button_upload.grid(row=3, column=0, padx=5, pady=5)
#############################################
#filter done by kim & adam

subj_options = ["","geography","math", "chinese","english","history","physics",'chemistry', 'biology', "computing", 'malay', 'tamil', 'hindi']
clicked = StringVar() # datatype of menu text
clicked.set( "--" ) # initial menu text

level_options = ["","p1", "p2", "p3", "p4", "p5", "p6", 's1', 's2', 's3', 's4']
clickedL = StringVar() # datatype of menu text
clickedL.set( "--" ) # initial menu text

# Create Dropdown menu
drop = ttk.OptionMenu( frame_button , clicked , *subj_options )
drop.grid(row=4, column=0, padx=5, pady=5)

dropL = ttk.OptionMenu( frame_button , clickedL , *level_options )
dropL.grid(row=5, column=0, padx=5, pady=5)

#for filter #done by zack
def filternote(): #ine 638, the clickedL is being read as PY_VAR1 , so we have to debug, most likely clicked also got problem
    filtered = tk.Tk()
    filtered.wm_title('filtered notes')
    for i in range(2,row_count+1):
        a = "A" + str(i)
        b = "B" + str(i)
        c = "C" + str(i)
        d = "D" + str(i)
        def filtered_notes():
            text_edit.delete(1.0, tk.END)
            text_edit.insert(tk.END, ws[f].value)
        if ws[c].value == clickedL and ws[b].value == clicked and ws[d].value == "public": #then display something idk
            print("filtered")
            liked_1 = ttk.Button(liked, text = ws[a].value, command=filtered_notes)
            liked_1.grid(row=i, column=0, padx=10, pady=10)
            liked_i.config(command=open_liked_note)

            
# filter button
button = ttk.Button(frame_button , text = "filter", command=filternote)
button.grid(row=6, column =0, padx=5, pady=5)
#############################################
#liked notes

        
def like_n(): #done by zack
    liked = tk.Tk()
    liked.wm_title('my liked notes')
    liked.geometry('700x700')
    for i in range(2,row_count+1):
        a = "A" + str(i)
        e = "E" + str(i)
        f = "F" + str(i)
        def open_liked_note():
            text_edit.delete(1.0, tk.END)
            text_edit.insert(tk.END, ws[f].value)
        if ws[e].value == "liked": #then display something idk
            print("show liked note")
        liked_i = ttk.Button(liked, text = ws[a].value, command=open_liked_note)
        liked_i.grid(row=i, column=0, padx=10, pady=10)
        liked_i.config(command=open_liked_note)
    


button_liked.config(command=like_n)

root.mainloop()
